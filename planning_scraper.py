#!/usr/bin/env python3
"""
planning_scraper.py  v5.0
=========================
Scrapes UK council Idox planning portals for warehouse, B8 and industrial
applications across North West England and Yorkshire.

HOW IT WORKS
------------
Uses Idox's simple keyword search URL — the most reliable method across all
Idox portal versions. Instead of navigating to the advanced form and clicking
buttons (which varies per portal), we pass the keyword directly in the URL:

  /online-applications/search.do?action=simple&searchType=Application&searchTerm=warehouse

Each keyword runs as a separate search. Results are parsed, date-filtered,
and deduplicated. The real applicationDetails URL is captured from each
result's link, giving working clickable links.

KEYWORDS
--------
Precise terms chosen to match only industrial/B8 planning descriptions.
Short single words like "storage" are avoided as they match residential
descriptions ("storage heater", "storage cupboard"). Instead we use terms
that only appear in commercial planning applications.

Usage
-----
    python planning_scraper.py                         # last 60 days, all councils
    python planning_scraper.py --days 30
    python planning_scraper.py --date-from 2026-03-01
    python planning_scraper.py --councils Bolton Wigan Leeds
    python planning_scraper.py --output april_leads.xlsx
    python planning_scraper.py --no-headless

Requirements
------------
    pip3 install playwright pandas openpyxl beautifulsoup4 lxml
    playwright install chromium
"""

import argparse
import os
import re
import sys
import time
from datetime import date, datetime, timedelta
from urllib.parse import quote_plus, urljoin

import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout

# ---------------------------------------------------------------------------
# Keywords — each searched separately via Idox simple search URL.
# Chosen to be specific enough that they almost never appear in residential,
# retail or agricultural planning descriptions.
# ---------------------------------------------------------------------------
DEFAULT_KEYWORDS = [
    "warehouse",
    "distribution centre",
    "logistics",
    "industrial unit",
    "storage and distribution",
    "employment land",
    "trade counter",
    "manufacturing",
    "use class b8",
    "b8 use class",
    "use class b2",
]

# Descriptions containing any of these phrases are excluded even if a keyword matched.
# Guards against edge cases like "warehouse conversion to flats".
_EXCLUDE_PHRASES = [
    "conversion to residential",
    "conversion to dwellings",
    "conversion to flats",
    "change of use to residential",
    "change of use to c3",
    "demolition of warehouse",          # demolition only — not relevant for new build
]

# ---------------------------------------------------------------------------
# Council registry — verified URLs only
# ---------------------------------------------------------------------------
# portal_type:
#   'idox'  — standard Idox Public Access simple keyword search
#   'skip'  — portal not automatable; manual URL provided

COUNCILS = [
    # ── North West ────────────────────────────────────────────────────────
    {
        # Confirmed: paplanning.bolton.gov.uk (not www.bolton.gov.uk which 404s)
        "name":        "Bolton",
        "portal_type": "idox",
        "base_url":    "https://paplanning.bolton.gov.uk",
        "search_url":  "https://paplanning.bolton.gov.uk/online-applications/search.do",
        "manual_url":  "https://paplanning.bolton.gov.uk/online-applications/",
    },
    {
        "name":        "Wigan",
        "portal_type": "idox",
        "base_url":    "https://planning.wigan.gov.uk",
        "search_url":  "https://planning.wigan.gov.uk/online-applications/search.do",
        "manual_url":  "https://planning.wigan.gov.uk/online-applications/",
    },
    {
        # Warrington uses SwiftLG with CSRF — not automatable
        "name":        "Warrington",
        "portal_type": "skip",
        "manual_url":  "https://planning.warrington.gov.uk/swiftlg/apas/run/wphappcriteria.display",
        "skip_reason": "SwiftLG portal — search manually at the link above",
    },
    {
        # Blackburn uses Northgate — times out, not reachable
        "name":        "Blackburn with Darwen",
        "portal_type": "skip",
        "manual_url":  "https://planning.blackburn.gov.uk/Northgate/PlanningExplorer/GeneralSearch.aspx",
        "skip_reason": "Northgate portal — search manually at the link above",
    },
    {
        "name":        "South Ribble",
        "portal_type": "idox",
        "base_url":    "https://publicaccess.southribble.gov.uk",
        "search_url":  "https://publicaccess.southribble.gov.uk/online-applications/search.do",
        "manual_url":  "https://publicaccess.southribble.gov.uk/online-applications/",
    },
    {
        # Preston refuses automated connections
        "name":        "Preston",
        "portal_type": "skip",
        "manual_url":  "https://publicaccess.preston.gov.uk/online-applications/",
        "skip_reason": "Server blocks automated access — search manually at the link above",
    },
    {
        "name":        "Rochdale",
        "portal_type": "idox",
        "base_url":    "https://publicaccess.rochdale.gov.uk",
        "search_url":  "https://publicaccess.rochdale.gov.uk/online-applications/search.do",
        "manual_url":  "https://publicaccess.rochdale.gov.uk/online-applications/",
    },
    {
        "name":        "Bury",
        "portal_type": "idox",
        "base_url":    "https://planning.bury.gov.uk",
        "search_url":  "https://planning.bury.gov.uk/online-applications/search.do",
        "manual_url":  "https://planning.bury.gov.uk/online-applications/",
    },
    {
        "name":        "Salford",
        "portal_type": "idox",
        "base_url":    "https://publicaccess.salford.gov.uk",
        "search_url":  "https://publicaccess.salford.gov.uk/online-applications/search.do",
        "manual_url":  "https://publicaccess.salford.gov.uk/online-applications/",
    },
    {
        "name":        "Trafford",
        "portal_type": "idox",
        "base_url":    "https://pa.trafford.gov.uk",
        "search_url":  "https://pa.trafford.gov.uk/online-applications/search.do",
        "manual_url":  "https://pa.trafford.gov.uk/online-applications/",
    },
    {
        "name":        "Manchester City",
        "portal_type": "idox",
        "base_url":    "https://pa.manchester.gov.uk",
        "search_url":  "https://pa.manchester.gov.uk/online-applications/search.do",
        "manual_url":  "https://pa.manchester.gov.uk/online-applications/",
    },
    {
        "name":        "St Helens",
        "portal_type": "idox",
        "base_url":    "https://publicaccess.sthelens.gov.uk",
        "search_url":  "https://publicaccess.sthelens.gov.uk/online-applications/search.do",
        "manual_url":  "https://publicaccess.sthelens.gov.uk/online-applications/",
    },
    # ── Yorkshire ─────────────────────────────────────────────────────────
    {
        "name":        "Leeds",
        "portal_type": "idox",
        "base_url":    "https://publicaccess.leeds.gov.uk",
        "search_url":  "https://publicaccess.leeds.gov.uk/online-applications/search.do",
        "manual_url":  "https://publicaccess.leeds.gov.uk/online-applications/",
    },
    {
        "name":        "Wakefield",
        "portal_type": "idox",
        "base_url":    "https://planning.wakefield.gov.uk",
        "search_url":  "https://planning.wakefield.gov.uk/online-applications/search.do",
        "manual_url":  "https://planning.wakefield.gov.uk/online-applications/",
    },
    {
        # Kirklees uses their own bespoke system at kirklees.gov.uk/beta/planning-applications/
        # Not an Idox portal — search.do does not exist here
        "name":        "Kirklees",
        "portal_type": "skip",
        "manual_url":  "https://www.kirklees.gov.uk/beta/planning-applications/search-for-planning-applications/default.aspx",
        "skip_reason": "Bespoke portal (not Idox) — search manually at the link above",
    },
    {
        "name":        "Bradford",
        "portal_type": "idox",
        "base_url":    "https://planning.bradford.gov.uk",
        "search_url":  "https://planning.bradford.gov.uk/online-applications/search.do",
        "manual_url":  "https://planning.bradford.gov.uk/online-applications/",
    },
    {
        "name":        "Sheffield",
        "portal_type": "idox",
        "base_url":    "https://planningapps.sheffield.gov.uk",
        "search_url":  "https://planningapps.sheffield.gov.uk/online-applications/search.do",
        "manual_url":  "https://planningapps.sheffield.gov.uk/online-applications/",
    },
    {
        # Rotherham uses Fastweb with no date filter — returns archive records from 1948
        "name":        "Rotherham",
        "portal_type": "skip",
        "manual_url":  "https://planning.rotherham.gov.uk/search.asp",
        "skip_reason": "Fastweb portal with no date validation — search manually at the link above",
    },
    {
        "name":        "Barnsley",
        "portal_type": "idox",
        "base_url":    "https://www.barnsley.gov.uk",
        "search_url":  "https://www.barnsley.gov.uk/online-applications/search.do",
        "manual_url":  "https://www.barnsley.gov.uk/online-applications/",
    },
    {
        # Confirmed: planning.doncaster.gov.uk (not doncaster.gov.uk which 404s)
        "name":        "Doncaster",
        "portal_type": "idox",
        "base_url":    "https://planning.doncaster.gov.uk",
        "search_url":  "https://planning.doncaster.gov.uk/online-applications/search.do",
        "manual_url":  "https://planning.doncaster.gov.uk/online-applications/",
    },
    {
        # Confirmed URL: newplanningaccess.eastriding.gov.uk
        # BUT: portal explicitly states "Automated search processes will be blocked"
        # Marking as skip to avoid being blocked
        "name":        "East Yorkshire",
        "portal_type": "skip",
        "manual_url":  "https://newplanningaccess.eastriding.gov.uk/newplanningaccess/search.do?action=simple&searchType=Application",
        "skip_reason": "Portal explicitly blocks automated access — search manually at the link above",
    },
    {
        # North Yorkshire: unverified — mark as skip until URL confirmed
        "name":        "North Yorkshire",
        "portal_type": "skip",
        "manual_url":  "https://www.northyorks.gov.uk/planning-and-building-control",
        "skip_reason": "Portal URL unconfirmed — search manually at the link above",
    },
]

# ---------------------------------------------------------------------------
# Parsing helpers
# ---------------------------------------------------------------------------

# Date patterns — ordered from most to least specific.
# Bolton (and some other councils) prefix dates with day-of-week: "Thu 29 Jan 2026"
# so patterns must optionally consume 3-letter day abbreviations.
_DAY_PREFIX = r'(?:Mon|Tue|Wed|Thu|Fri|Sat|Sun)\s+'
_DATE_PATTERNS = [
    # "Validated: Thu 29 Jan 2026"  or  "Validated: 29 Jan 2026"
    (r'[Vv]alidated[:\s]+(?:' + _DAY_PREFIX + r')?(\d{1,2}\s+\w+\s+\d{4})',
     ['%d %b %Y', '%d %B %Y']),
    # "Validated: 29/01/2026"
    (r'[Vv]alidated[:\s]+(\d{1,2}/\d{1,2}/\d{4})',
     ['%d/%m/%Y']),
    # "Date Validated: Thu 29 Jan 2026"
    (r'Date\s+[Vv]alidated[:\s]+(?:' + _DAY_PREFIX + r')?(\d{1,2}\s+\w+\s+\d{4})',
     ['%d %b %Y', '%d %B %Y']),
    # "Date Validated: 29/01/2026"
    (r'Date\s+[Vv]alidated[:\s]+(\d{1,2}/\d{1,2}/\d{4})',
     ['%d/%m/%Y']),
    # "Received: Thu 29 Jan 2026"  — fallback when validated not present
    (r'[Rr]eceived[:\s]+(?:' + _DAY_PREFIX + r')?(\d{1,2}\s+\w+\s+\d{4})',
     ['%d %b %Y', '%d %B %Y']),
    # "Received: 29/01/2026"
    (r'[Rr]eceived[:\s]+(\d{1,2}/\d{1,2}/\d{4})',
     ['%d/%m/%Y']),
]


def _parse_validated_date(text: str):
    """Extract a validated (or received) date from Idox metaInfo text."""
    # Collapse whitespace so multi-line Bolton-style text is handled cleanly
    text = re.sub(r'\s+', ' ', text)
    for pattern, fmts in _DATE_PATTERNS:
        m = re.search(pattern, text)
        if m:
            raw = m.group(1).strip()
            for fmt in fmts:
                try:
                    return datetime.strptime(raw, fmt).date()
                except ValueError:
                    continue
    return None


def _parse_ref(text: str) -> str:
    text = re.sub(r'\s+', ' ', text)
    m = re.search(r'Ref\.?\s*No[.:\s]+([A-Z0-9/_-]+)', text, re.IGNORECASE)
    return m.group(1).strip() if m else ""


def _parse_status(item_soup) -> str:
    s = item_soup.find('span', class_='status')
    if s:
        return s.get_text(strip=True)
    meta = item_soup.find('p', class_='metaInfo')
    if meta:
        text = re.sub(r'\s+', ' ', meta.get_text())
        m = re.search(r'Status[:\s]+(.+?)(?:\||$)', text)
        if m:
            return m.group(1).strip()
    return ""


def _is_excluded(desc: str) -> bool:
    """Return True if description contains an exclusion phrase."""
    dl = desc.lower()
    return any(phrase in dl for phrase in _EXCLUDE_PHRASES)


def _get_next_page_url(html: str, base_url: str):
    soup = BeautifulSoup(html, 'lxml')
    nxt = (
        soup.find('a', class_='next')
        or soup.find('a', string=re.compile(r'^\s*Next\s*$', re.I))
        or soup.find('a', title=re.compile(r'Next', re.I))
    )
    if nxt and nxt.get('href'):
        return urljoin(base_url, nxt['href'])
    return None


def _parse_idox_results(html: str, base_url: str, council_name: str,
                         date_from=None, date_to=None) -> list:
    """
    Parse one page of Idox search results.
    Handles both standard Idox (description in <span class='description'>)
    and Bolton-style Idox (description in the link text / <a> title attribute).
    """
    soup    = BeautifulSoup(html, 'lxml')
    results = []

    for item in soup.select('li.searchresult'):
        link_el   = item.find('a')
        raw_href  = link_el.get('href', '') if link_el else ''
        full_link = urljoin(base_url, raw_href) if raw_href else ''

        address_el = item.find('span', class_='address')
        desc_el    = item.find('span', class_='description')
        meta_el    = item.find('p', class_='metaInfo')

        address = address_el.get_text(strip=True) if address_el else ''

        # Description: try <span class="description"> first, then <a> title attribute.
        # Bolton-style Idox puts proposal in title= and address as link body text.
        if desc_el:
            description = desc_el.get_text(strip=True)
        elif link_el:
            title_attr  = link_el.get('title', '').strip()
            link_text   = link_el.get_text(strip=True)
            if title_attr:
                description = title_attr
                # If address wasn't in a span, the link body text IS the address
                if not address:
                    address = link_text
            else:
                description = link_text
        else:
            description = ''

        meta_text = meta_el.get_text(separator=' ', strip=True) if meta_el else ''

        # Skip if description matches an exclusion phrase
        if _is_excluded(description):
            continue

        validated = _parse_validated_date(meta_text)

        # Date filter
        if date_from and validated and validated < date_from:
            continue
        if date_to and validated and validated > date_to:
            continue
        # Drop undated results when a date window is active
        if (date_from or date_to) and not validated:
            continue

        results.append({
            'Council':     council_name,
            'Reference':   _parse_ref(meta_text),
            'Address':     address,
            'Description': description,
            'Status':      _parse_status(item),
            'Validated':   validated.isoformat() if validated else '',
            'Link':        full_link,
        })

    return results


def _fetch_applicant(page, link_url: str) -> str:
    """
    Fetch the application detail contacts tab and extract applicant name.
    Returns empty string if not found or fetch fails.
    """
    if not link_url:
        return ''
    try:
        contacts_url = link_url.replace('activeTab=summary', 'activeTab=contacts')
        if 'activeTab=' not in contacts_url:
            contacts_url = link_url + ('&' if '?' in link_url else '?') + 'activeTab=contacts'
        page.goto(contacts_url, timeout=20000, wait_until='domcontentloaded')
        time.sleep(0.5)
        soup = BeautifulSoup(page.content(), 'lxml')
        for row in soup.select('tr'):
            cells = row.find_all(['th', 'td'])
            if len(cells) >= 2:
                header = cells[0].get_text(strip=True).lower()
                if 'applicant' in header and 'name' in header:
                    val = cells[1].get_text(strip=True)
                    if val:
                        return val
    except Exception:
        pass
    return ''


# ---------------------------------------------------------------------------
# Retry — abort immediately on DNS failures
# ---------------------------------------------------------------------------

_HARD_FAIL = ('ERR_NAME_NOT_RESOLVED', 'ERR_ADDRESS_UNREACHABLE',
              'Name or service not known', 'getaddrinfo')


def _retry(fn, retries: int = 3, delay: float = 4.0, label: str = ""):
    last_exc = None
    for attempt in range(1, retries + 1):
        try:
            return fn()
        except Exception as exc:
            last_exc = exc
            if any(e in str(exc) for e in _HARD_FAIL):
                raise RuntimeError(
                    f"[{label}] DNS failure — URL does not exist: {exc}"
                ) from exc
            print(f"  [{label}] Attempt {attempt}/{retries} failed: {exc}")
            if attempt < retries:
                time.sleep(delay * attempt)
    raise RuntimeError(f"[{label}] All {retries} attempts failed. Last: {last_exc}")


# ---------------------------------------------------------------------------
# Idox scraper — keyword search with GET-then-form-POST fallback
# ---------------------------------------------------------------------------

def _build_keyword_url(search_url: str, keyword: str) -> str:
    """Build an Idox simple search URL with the keyword as a GET parameter."""
    kw_enc = quote_plus(keyword)
    base   = search_url.split('?')[0]
    return f"{base}?action=simple&searchType=Application&searchTerm={kw_enc}"


def _has_results(html: str) -> bool:
    """Return True if the page contains actual search results."""
    return 'li class="searchresult"' in html or "li class='searchresult'" in html


def _looks_like_empty_form(html: str) -> bool:
    """
    Return True if the page loaded but shows an empty search form —
    meaning the GET keyword parameter was ignored by this Idox version.
    """
    has_form    = 'searchTerm' in html or 'Simple Search' in html
    has_results = _has_results(html)
    has_none    = 'noResults' in html or 'No results' in html
    return has_form and not has_results and not has_none


def _make_page(ctx):
    """Create a fresh page, closing any existing ones first."""
    for p in ctx.pages:
        try:
            p.close()
        except Exception:
            pass
    return ctx.new_page()


def _submit_search_form(page, search_url: str, keyword: str) -> str:
    """
    Navigate to the Idox simple search page, fill the keyword into the
    search box, submit via Enter key (most reliable cross-version method),
    and return the resulting page HTML.
    """
    base = search_url.split('?')[0]
    form_url = f"{base}?action=simple&searchType=Application"
    page.goto(form_url, timeout=45000, wait_until='domcontentloaded')
    time.sleep(1.5)  # let the page fully settle

    # Find and fill the search input
    filled = False
    input_selector = None
    for selector in [
        'input#simpleSearchString',
        'input[name="searchTerm"]',
        'input[id*="search"]',
        'input[type="text"]',
    ]:
        try:
            page.wait_for_selector(selector, timeout=3000)
            page.fill(selector, keyword, timeout=3000)
            input_selector = selector
            filled = True
            break
        except Exception:
            continue

    if not filled:
        return page.content()

    time.sleep(0.3)

    # Submit — try Enter key first (most reliable), then button click
    submitted = False
    if input_selector:
        try:
            page.press(input_selector, 'Enter')
            submitted = True
        except Exception:
            pass

    if not submitted:
        for selector in [
            'input[type="submit"][value="Search"]',
            'input[type="submit"]',
            'button[type="submit"]',
            'button:has-text("Search")',
        ]:
            try:
                page.click(selector, timeout=3000)
                submitted = True
                break
            except Exception:
                continue

    # Wait for navigation and results
    try:
        page.wait_for_load_state('networkidle', timeout=20000)
    except PWTimeout:
        pass
    try:
        page.wait_for_selector(
            'li.searchresult, p.noResults, .no-results, #noResultsMessage',
            timeout=15000
        )
    except PWTimeout:
        pass
    time.sleep(1.0)
    return page.content()


def scrape_idox_council(ctx, council: dict, keywords: list,
                         date_from=None, date_to=None, page_delay=1.5,
                         debug=False, fetch_applicants=False) -> list:
    """
    For each keyword, search the Idox portal.
    First tries GET with searchTerm in the URL (works on most Idox versions).
    If that returns an empty form (some councils ignore GET params), falls back
    to filling and submitting the form via Playwright.
    Pass debug=True to save raw HTML to /tmp/<council>_debug.html.
    """
    all_results = []
    seen_refs   = set()
    name        = council['name']
    base_url    = council['base_url']
    search_url  = council['search_url']

    # Fresh page for this council
    page = _make_page(ctx)

    # Detect on first keyword whether this council needs form submission
    needs_form_submit = False

    for kw in keywords:
        url      = _build_keyword_url(search_url, kw)
        page_num = 1

        def _load(u=url, kword=kw):
            nonlocal page, needs_form_submit
            try:
                page.url
            except Exception:
                page = _make_page(ctx)

            # Try GET approach first
            page.goto(u, timeout=45000, wait_until='domcontentloaded')
            try:
                page.wait_for_selector(
                    'li.searchresult, p.noResults, .no-results, #noResultsMessage',
                    timeout=8000
                )
            except PWTimeout:
                pass
            time.sleep(0.8)
            html = page.content()

            # If GET returned an empty form, switch to form-submit mode
            if _looks_like_empty_form(html):
                needs_form_submit = True
                html = _submit_search_form(page, search_url, kword)

            return html

        try:
            html = _retry(_load, retries=3, delay=4.0, label=f"{name}/{kw}")
        except RuntimeError as exc:
            print(f"  [{name}] SKIPPING '{kw}': {exc}")
            continue

        # Debug mode: save raw HTML so you can inspect what the scraper received
        if debug and kw == keywords[0]:
            debug_path = f"/tmp/{name.replace(' ','_')}_debug.html"
            with open(debug_path, 'w', encoding='utf-8') as f:
                f.write(html)
            print(f"  [{name}] DEBUG: raw HTML saved to {debug_path}")
            print(f"  [{name}] DEBUG: page length={len(html)}, "
                  f"searchresults={'YES' if _has_results(html) else 'NO'}, "
                  f"empty_form={'YES' if _looks_like_empty_form(html) else 'NO'}, "
                  f"noResults={'YES' if 'noResults' in html else 'NO'}")

        kw_count = 0
        while html:
            batch = _parse_idox_results(html, base_url, name, date_from, date_to)

            for r in batch:
                key = r['Reference'] or r['Link']
                if key and key not in seen_refs:
                    seen_refs.add(key)
                    all_results.append(r)
                    kw_count += 1

            next_url = _get_next_page_url(html, base_url)
            if not next_url or next_url == page.url:
                break
            page_num += 1
            time.sleep(page_delay)

            def _next(u=next_url):
                nonlocal page
                try:
                    page.url
                except Exception:
                    page = _make_page(ctx)
                page.goto(u, timeout=30000, wait_until='domcontentloaded')
                try:
                    page.wait_for_selector('li.searchresult, p.noResults', timeout=12000)
                except PWTimeout:
                    pass
                time.sleep(0.5)
                return page.content()

            try:
                html = _retry(_next, retries=2, delay=2.0, label=f"{name}/{kw}/p{page_num}")
            except RuntimeError:
                break

        if kw_count:
            mode = " [form]" if needs_form_submit else ""
            print(f"  [{name}] '{kw}'{mode}: {kw_count} new results")

    # Optionally fetch applicant name from each application's detail page
    if fetch_applicants and all_results:
        print(f"  [{name}] Fetching applicant names ({len(all_results)} applications)...")
        for r in all_results:
            r['Applicant'] = _fetch_applicant(page, r.get('Link', ''))
            time.sleep(0.5)  # be polite to the server

    return all_results


# ---------------------------------------------------------------------------
# Dispatcher
# ---------------------------------------------------------------------------

def scrape_council(ctx, council: dict, keywords: list,
                   date_from=None, date_to=None, page_delay=1.5,
                   debug=False, fetch_applicants=False) -> list:
    pt = council.get('portal_type', 'idox')
    if pt == 'idox':
        return scrape_idox_council(ctx, council, keywords, date_from, date_to,
                                   page_delay, debug=debug,
                                   fetch_applicants=fetch_applicants)
    elif pt == 'skip':
        reason = council.get('skip_reason', 'Portal not automatable.')
        print(f"  [{council['name']}] SKIPPED — {reason}")
        print(f"    Manual search: {council['manual_url']}")
        return []
    else:
        print(f"  [{council['name']}] Unknown portal type '{pt}' — skipping.")
        return []


# ---------------------------------------------------------------------------
# Deduplication across councils
# ---------------------------------------------------------------------------

def deduplicate(results: list) -> list:
    seen, out = set(), []
    for r in results:
        key = (r['Council'], r['Reference'] or r['Link'])
        if key not in seen:
            seen.add(key)
            out.append(r)
    return out


# ---------------------------------------------------------------------------
# Excel output with working hyperlinks
# ---------------------------------------------------------------------------

_ORANGE = "E8620A"
_WHITE  = "FFFFFF"
_DARK   = "1A1A1A"
_LIGHT  = "FFF5EE"
_MID    = "FFE4D0"
_THIN   = Side(style='thin', color='DDDDDD')
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_STATUS_FILL = {
    'approved':              'C8E6C9',
    'application approved':  'C8E6C9',
    'permitted':             'C8E6C9',
    'grant':                 'C8E6C9',
    'pending':               'FFF9C4',
    'pending consideration': 'FFF9C4',
    'under consideration':   'FFF9C4',
    'refused':               'FFCDD2',
    'withdrawn':             'F5F5F5',
}


def write_excel(data: list, output_path: str, date_from=None, date_to=None):
    wb = Workbook()

    # ── Planning Leads sheet ─────────────────────────────────────────────
    ws = wb.active
    ws.title = "Planning Leads"
    ws.sheet_view.showGridLines = False

    # Check if any applicant data was fetched
    has_applicant = any(r.get('Applicant') for r in data)

    if has_applicant:
        headers    = ['Council', 'Reference', 'Address', 'Applicant', 'Description', 'Status', 'Validated', 'View Application']
        col_widths = [18, 22, 36, 28, 55, 24, 13, 18]
        data_keys  = ['Council', 'Reference', 'Address', 'Applicant', 'Description', 'Status', 'Validated']
        desc_col   = 5  # Description column for wrap_text
        link_col   = 8
    else:
        headers    = ['Council', 'Reference', 'Address', 'Description', 'Status', 'Validated', 'View Application']
        col_widths = [18, 22, 38, 60, 26, 13, 18]
        data_keys  = ['Council', 'Reference', 'Address', 'Description', 'Status', 'Validated']
        desc_col   = 4
        link_col   = 7

    status_col = data_keys.index('Status') + 1

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = Font(name='Arial', bold=True, color=_WHITE, size=10)
        cell.fill      = PatternFill('solid', fgColor=_ORANGE)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = _BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 22
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    for row_idx, row in enumerate(data, 2):
        bg = _MID if row_idx % 2 == 0 else _LIGHT

        for col_idx, key in enumerate(data_keys, 1):
            val  = row.get(key, '') or ''
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font      = Font(name='Arial', size=9)
            cell.fill      = PatternFill('solid', fgColor=bg)
            # Wrap text on description and address columns
            cell.alignment = Alignment(vertical='top',
                                       wrap_text=(col_idx in (desc_col, 3)))
            cell.border    = _BORDER

        # Status colour coding
        status_val  = (row.get('Status') or '').lower()
        status_cell = ws.cell(row=row_idx, column=status_col)
        for key, colour in _STATUS_FILL.items():
            if key in status_val:
                status_cell.fill = PatternFill('solid', fgColor=colour)
                break

        # Hyperlink — real applicationDetails.do URL from portal
        link_cell = ws.cell(row=row_idx, column=link_col)
        link_url  = row.get('Link', '') or ''
        if link_url and link_url.startswith('http'):
            link_cell.value     = "Open in Portal"
            link_cell.hyperlink = link_url
            link_cell.font      = Font(name='Arial', size=9, color='0070C0', underline='single')
        else:
            link_cell.value = "—"
            link_cell.font  = Font(name='Arial', size=9, color='888888')
        link_cell.fill      = PatternFill('solid', fgColor=bg)
        link_cell.alignment = Alignment(horizontal='center', vertical='top')
        link_cell.border    = _BORDER

        ws.row_dimensions[row_idx].height = 45

    # ── Summary sheet ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions['A'].width = 34
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['D'].width = 28
    ws2.column_dimensions['E'].width = 12

    def _h1(r, c, v):
        cell = ws2.cell(r, c, v)
        cell.font = Font(name='Arial', bold=True, size=14, color=_ORANGE)

    def _label(r, c, v):
        cell = ws2.cell(r, c, v)
        cell.font = Font(name='Arial', bold=True, size=10)

    def _val(r, c, v):
        cell = ws2.cell(r, c, v)
        cell.font = Font(name='Arial', size=10)

    _h1(1, 1, 'Palletower — Planning Lead Intelligence')
    _val(2, 1, f'Generated: {date.today().strftime("%d %B %Y")}')
    ws2.cell(2, 1).font = Font(name='Arial', size=10, color='888888')

    window_str = ''
    if date_from:
        window_str = f"From {date_from.strftime('%d %b %Y')}"
    if date_to:
        window_str += f" to {date_to.strftime('%d %b %Y')}"
    if window_str:
        _val(3, 1, window_str)
        ws2.cell(3, 1).font = Font(name='Arial', size=10, color='888888')

    _label(5, 1, 'Total leads')
    ws2.cell(5, 2, len(data)).font = Font(name='Arial', bold=True, size=12, color=_ORANGE)

    # By council
    _label(7, 1, 'Council')
    _label(7, 2, 'Leads')
    for c in [ws2.cell(7, 1), ws2.cell(7, 2)]:
        c.fill   = PatternFill('solid', fgColor=_ORANGE)
        c.font   = Font(name='Arial', bold=True, size=10, color=_WHITE)
        c.border = _BORDER

    council_counts: dict = {}
    for r in data:
        council_counts[r['Council']] = council_counts.get(r['Council'], 0) + 1

    for i, (cname, cnt) in enumerate(sorted(council_counts.items()), 8):
        bg = _MID if i % 2 == 0 else _LIGHT
        for col, val in [(1, cname), (2, cnt)]:
            c = ws2.cell(i, col, val)
            c.font   = Font(name='Arial', size=10)
            c.fill   = PatternFill('solid', fgColor=bg)
            c.border = _BORDER

    # By status
    _label(7, 4, 'Status')
    _label(7, 5, 'Count')
    for c in [ws2.cell(7, 4), ws2.cell(7, 5)]:
        c.fill   = PatternFill('solid', fgColor=_ORANGE)
        c.font   = Font(name='Arial', bold=True, size=10, color=_WHITE)
        c.border = _BORDER

    status_counts: dict = {}
    for r in data:
        s = r.get('Status') or 'Unknown'
        status_counts[s] = status_counts.get(s, 0) + 1

    for i, (s, cnt) in enumerate(sorted(status_counts.items()), 8):
        bg = _MID if i % 2 == 0 else _LIGHT
        for col, val in [(4, s), (5, cnt)]:
            c = ws2.cell(i, col, val)
            c.font   = Font(name='Arial', size=10)
            c.fill   = PatternFill('solid', fgColor=bg)
            c.border = _BORDER

    wb.save(output_path)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _parse_date_arg(s: str) -> date:
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise argparse.ArgumentTypeError(
        f"Cannot parse date '{s}'. Use YYYY-MM-DD or DD/MM/YYYY."
    )


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description='Scrape Idox planning portals for warehouse/B8 leads — Palletower.',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    p.add_argument('--date-from', type=_parse_date_arg, default=None, metavar='DATE',
                   help='Only include results validated on or after DATE (YYYY-MM-DD or DD/MM/YYYY)')
    p.add_argument('--date-to',   type=_parse_date_arg, default=None, metavar='DATE',
                   help='Only include results validated on or before DATE')
    p.add_argument('--days',      type=int, default=60, metavar='N',
                   help='Look back N days from today (default 60). Ignored if --date-from set.')
    p.add_argument('--all-dates', action='store_true', default=False,
                   help='No date filter — return all matching applications (slow, noisy).')
    p.add_argument('--output',    default='leads.xlsx', metavar='FILE',
                   help='Output Excel file (default: leads.xlsx)')
    p.add_argument('--keywords',  nargs='+', default=DEFAULT_KEYWORDS, metavar='KW',
                   help='Keywords to search for (space-separated; quote multi-word terms)')
    p.add_argument('--councils',  nargs='+', default=None, metavar='COUNCIL',
                   help='Restrict to named councils, e.g. --councils Bolton Wigan Leeds')
    p.add_argument('--headless',  action=argparse.BooleanOptionalAction, default=True,
                   help='Run headless (default). Use --no-headless to watch the browser.')
    p.add_argument('--page-delay', type=float, default=1.5, metavar='SECS',
                   help='Pause between paginated requests (default 1.5s).')
    p.add_argument('--debug', action='store_true', default=False,
                   help='Save raw HTML from first keyword per council to /tmp/<council>_debug.html')
    p.add_argument('--fetch-applicants', action='store_true', default=False,
                   help='Fetch applicant name from each application detail page (slower — one extra request per result)')
    return p


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def run(args=None):
    parser = build_parser()
    args   = parser.parse_args(args)

    # Resolve date window
    if args.all_dates:
        args.date_from = None
        args.date_to   = None
    elif args.date_from is None and args.date_to is None:
        args.date_from = date.today() - timedelta(days=args.days)

    # Ensure .xlsx extension
    output = args.output
    if not output.endswith('.xlsx'):
        output = output.rsplit('.', 1)[0] + '.xlsx'

    # Council filter
    councils = COUNCILS
    if args.councils:
        requested = {c.lower() for c in args.councils}
        councils  = [c for c in COUNCILS if c['name'].lower() in requested]
        if not councils:
            print(f"ERROR: No councils matched. Available:")
            for c in COUNCILS:
                print(f"  {c['name']}")
            sys.exit(1)

    print("=" * 62)
    print("Planning Portal Scraper  v5.0 — Palletower")
    print("=" * 62)
    if args.all_dates:
        print("  Date window : ALL (no filter)")
    else:
        window = (f"  (last {args.days} days)"
                  if not args.date_to and args.date_from == date.today() - timedelta(days=args.days)
                  else "")
        print(f"  Date from   : {args.date_from}{window}")
        if args.date_to:
            print(f"  Date to     : {args.date_to}")
    print(f"  Keywords    : {', '.join(args.keywords)}")
    print(f"  Councils    : {len(councils)}")
    print(f"  Output      : {output}")
    print(f"  Headless    : {args.headless}")
    print("=" * 62)

    all_results, errors = [], []

    with sync_playwright() as pw:
        browser = pw.chromium.launch(
            headless=args.headless,
            args=['--ignore-certificate-errors'],
        )
        ctx = browser.new_context(
            user_agent=(
                'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) '
                'AppleWebKit/537.36 (KHTML, like Gecko) '
                'Chrome/124.0.0.0 Safari/537.36'
            ),
            viewport={'width': 1280, 'height': 900},
            ignore_https_errors=True,
        )

        for i, council in enumerate(councils, 1):
            print(f"\n[{i}/{len(councils)}] {council['name']}")
            try:
                data = scrape_council(
                    ctx, council,
                    keywords=args.keywords,
                    date_from=args.date_from,
                    date_to=args.date_to,
                    page_delay=args.page_delay,
                    debug=args.debug,
                    fetch_applicants=args.fetch_applicants,
                )
                all_results.extend(data)
                if data:
                    print(f"  -> {len(data)} results")
                else:
                    print(f"  -> 0 results")
            except Exception as exc:
                errors.append(f"{council['name']}: {exc}")
                print(f"  ERROR — {exc}")

        browser.close()

    before      = len(all_results)
    all_results = deduplicate(all_results)
    print(f"\nDeduplication: {before} → {len(all_results)} unique results")

    # Sort most recent first
    all_results.sort(key=lambda r: r.get('Validated') or '0000-00-00', reverse=True)

    write_excel(all_results, output, date_from=args.date_from, date_to=args.date_to)
    abs_path = os.path.abspath(output)
    print(f"Saved {len(all_results)} leads → {abs_path}")

    print("\n" + "=" * 62)
    print("SUMMARY")
    print("=" * 62)
    if all_results:
        counts: dict = {}
        for r in all_results:
            counts[r['Council']] = counts.get(r['Council'], 0) + 1
        for cname, n in sorted(counts.items()):
            print(f"  {cname:<36} {n:>4} lead(s)")
        print(f"  {'TOTAL':<36} {len(all_results):>4}")
    else:
        print("  No results found.")

    # Manual search links for skipped councils
    skipped = [c for c in councils if c.get('portal_type') == 'skip']
    if skipped:
        print("\nSearch these councils manually:")
        for c in skipped:
            print(f"  {c['name']:<36} {c['manual_url']}")

    if errors:
        print(f"\nErrors ({len(errors)}):")
        for e in errors:
            print(f"  ! {e}")

    if all_results:
        if sys.platform == 'darwin':
            os.system(f'open "{abs_path}"')
        elif sys.platform == 'win32':
            os.startfile(abs_path)

    return all_results


if __name__ == '__main__':
    run()
# placeholder
